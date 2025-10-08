from collections import defaultdict

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from django.db.models import Q
from django.db.models.aggregates import Count, Avg
from django.db.models.functions import Round
from django.shortcuts import get_object_or_404
from django_filters.rest_framework import DjangoFilterBackend
from geopy.distance import geodesic
from openpyxl import load_workbook
from rest_framework import viewsets, status
from rest_framework.decorators import api_view
from rest_framework.filters import OrderingFilter
from rest_framework.filters import SearchFilter
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.views import APIView

from app_run.models import Run, AthleteInfo, Challenge, Positions, CollectibleItem, Subscribe, User
from app_run.serializers import RunSerializer, UserListSerializer, AthleteInfoSerializer, ChallengeSerializer, \
    PositionsSerializer, CollectibleItemSerializer, CoachDetailSerializer, AthleteDetailSerializer
from app_run.utils import award_challenge_if_completed_run_10, calculate_run_time_in_seconds
from .utils import calculate_and_save_run_distance, award_challenge_if_completed_run_50km, collect_item_if_nearby


@api_view(['GET'])
def company_details(request):
    return Response({
        'company_name': settings.COMPANY_NAME,
        'slogan': settings.SLOGAN,
        'contacts': settings.CONTACTS,
    })


class Pagination(PageNumberPagination):
    page_size_query_param = 'size'


class RunViewSet(viewsets.ModelViewSet):
    queryset = Run.objects.all().select_related('athlete')
    serializer_class = RunSerializer
    filter_backends = DjangoFilterBackend, OrderingFilter,
    filterset_fields = ('status', 'athlete')
    ordering_fields = ('created_at',)
    pagination_class = Pagination


class RunStarView(APIView):
    def post(self, request, id):
        run = get_object_or_404(Run, id=id)
        if run.status == Run.Status.INIT:
            run.status = Run.Status.IN_PROGRESS
            run.save()
            return Response({'message':
                                 'Run has started'},
                            status=status.HTTP_200_OK)
        return Response({'message':
                             f'The run status must be "init"; current status:{run.status}'},
                        status=status.HTTP_400_BAD_REQUEST)


class RunStopView(APIView):
    def post(self, request, id):
        run = get_object_or_404(
            Run.objects.annotate(
                speed_avg=Round(Avg('positions__speed'), 2)),
            id=id)
        if run.status == Run.Status.IN_PROGRESS:
            run.status = Run.Status.FINISHED
            run.run_time_seconds = calculate_run_time_in_seconds(run)
            run.speed = run.speed_avg
            run.save()
            award_challenge_if_completed_run_10(athlete_id=run.athlete.id)
            distance = calculate_and_save_run_distance(run_id=id)
            award_challenge_if_completed_run_50km(athlete_id=run.athlete.id)

            if distance >= 2 and run.run_time_seconds <= 600:
                Challenge.objects.create(athlete=run.athlete, full_name=Challenge.NameChoices.RUN2KMIN10M)

            return Response({'message':
                                 'Run has finished'},
                            status=status.HTTP_200_OK)
        return Response({'message':
                             f'The run status must be "in_process"; current status:{run.status}'},
                        status=status.HTTP_400_BAD_REQUEST)


class UserViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = (get_user_model().objects.
                exclude(is_superuser=True).
                annotate(runs_finished=Count('runs', filter=Q(runs__status=Run.Status.FINISHED)),
                         rating=Avg('followers__rating'), ))

    filter_backends = (SearchFilter, OrderingFilter)
    search_fields = 'first_name', 'last_name'
    ordering_fields = ('date_joined',)
    pagination_class = Pagination

    def get_queryset(self):
        qs = super().get_queryset()

        user_type = self.request.query_params.get('type', None)
        if user_type == 'coach':
            return qs.filter(is_staff=True)
        elif user_type == 'athlete':
            return qs.filter(is_staff=False)
        else:
            return qs

    def get_object(self):
        obj = super().get_object()
        if self.action == 'retrieve':
            self._is_staff = obj.is_staff
        return obj

    def get_serializer_class(self):
        if self.action == 'retrieve':
            if self._is_staff:
                return CoachDetailSerializer
            else:
                return AthleteDetailSerializer
        else:
            return UserListSerializer


class AthleteInfoView(APIView):

    def _get_user(self, pk):
        return get_object_or_404(get_user_model(), pk=pk)

    def get(self, request, user_id):
        user = self._get_user(user_id)
        athlete_info, _ = AthleteInfo.objects.get_or_create(athlete_id=user.id)
        return Response({'user_id': athlete_info.athlete.id,
                         'weight': athlete_info.weight,
                         'goals': athlete_info.goals})

    def put(self, request, user_id):
        user = self._get_user(user_id)
        serializer = AthleteInfoSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        athlete_info, _ = (AthleteInfo.objects.
                           update_or_create(athlete_id=user.id,
                                            defaults={
                                                'weight': serializer.validated_data.get('weight',
                                                                                        None),
                                                'goals': serializer.validated_data.get('goals',
                                                                                       None)}))

        return Response({'message':
                             'Athlete Info has created or updated'},
                        status=status.HTTP_201_CREATED)


class ChallengesView(viewsets.ReadOnlyModelViewSet):
    serializer_class = ChallengeSerializer

    def get_queryset(self):
        athlete_id = self.request.query_params.get('athlete', None)
        if athlete_id:
            return Challenge.objects.filter(athlete_id=athlete_id).select_related('athlete')
        return Challenge.objects.all().select_related('athlete')


class PositionsViewSet(viewsets.ModelViewSet):
    serializer_class = PositionsSerializer

    def perform_create(self, serializer):
        latitude = serializer.validated_data.get('latitude')
        longitude = serializer.validated_data.get('longitude')

        run = serializer.validated_data.get('run')
        previous_position = Positions.objects.filter(run=run).order_by('-id').first()

        position_distance = 0
        speed = 0
        if previous_position:
            distance_to_previous = round(geodesic((latitude, longitude),
                                                  (previous_position.latitude, previous_position.longitude)).kilometers,
                                         2)
            if previous_position.distance:
                position_distance = previous_position.distance + distance_to_previous
            else:
                position_distance = distance_to_previous

            time_from_previous = (
                    serializer.validated_data.get('date_time') - previous_position.date_time).total_seconds()

            if time_from_previous:
                speed = round(distance_to_previous * 1000 / time_from_previous, 2)

        instance = serializer.save(distance=position_distance,
                                   speed=speed)
        collect_item_if_nearby(
            latitude=latitude,
            longitude=longitude,
            user=instance.run.athlete
        )

    def get_queryset(self):
        qs = Positions.objects.all()
        run_id = self.request.query_params.get('run', None)
        if run_id:
            return qs.filter(run_id=run_id)
        return qs


@api_view(['POST'])
def upload_file(request):
    file = request.FILES.get('file')
    EXPECTED_HEADERS = ['Name', 'UID', 'Value', 'Latitude', 'Longitude', 'URL']
    wb = load_workbook(file)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    if headers != EXPECTED_HEADERS:
        return Response({
            'error': 'Wrong headers',
            'expected': EXPECTED_HEADERS,
            'got': headers,
        }, status=status.HTTP_400_BAD_REQUEST)

    invalid_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue
        data = {
            'name': row[0],
            'uid': row[1],
            'value': row[2],
            'latitude': row[3],
            'longitude': row[4],
            'picture': row[5]
        }
        serializer = CollectibleItemSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
        else:
            invalid_rows.append(row)
    return Response(invalid_rows)


class CollectibleItemViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = CollectibleItemSerializer
    queryset = CollectibleItem.objects.all()


@api_view(['GET'])
def challenge_summary(request):
    challenges = Challenge.objects.select_related('athlete').all()
    challenges_dict = defaultdict(list)
    for challenge in challenges:
        challenges_dict[challenge.get_full_name_display()].append({
            'id': challenge.athlete.id,
            'username': challenge.athlete.username,
            'full_name': challenge.athlete.get_full_name(),
        })

    data = []
    for key, value in challenges_dict.items():
        data.append({
            'name_to_display': key,
            'athletes': value
        })

    return Response(data, status=status.HTTP_200_OK)


@api_view(['POST'])
def subscribe_coach(request, coach_id):
    athlete_id = request.data.get('athlete', None)
    if athlete_id:
        try:
            athlete = get_user_model().objects.get(id=athlete_id, is_staff=False)
        except get_user_model().DoesNotExist:
            return Response({'message': 'Invalid athlete id'}, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response({'message': 'athlete field is required '}, status=status.HTTP_400_BAD_REQUEST)

    coach = get_object_or_404(User, id=coach_id)
    if not coach.is_staff:
        return Response(
            {'message': f'User is not a coach'},
            status=status.HTTP_400_BAD_REQUEST)

    try:
        Subscribe.objects.create(subscriber=athlete, subscribed_to=coach)
    except IntegrityError:
        return Response(
            {'message': f'Subscribe athlete with id {athlete.id} for coach with id {coach.id} already exists'},
            status=status.HTTP_400_BAD_REQUEST)
    return Response({f'message': f'Athlete with id {athlete.id} successfully subscribed coach with id {coach.id}'},
                    status=status.HTTP_200_OK)


@api_view(['POST'])
def rate_coach(request, coach_id):
    athlete_id = request.data.get('athlete', None)
    if not athlete_id:
        return Response({'message': 'athlete field is required '}, status=status.HTTP_400_BAD_REQUEST)

    try:
        athlete = get_user_model().objects.get(id=athlete_id, is_staff=False)
    except get_user_model().DoesNotExist:
        return Response({'message': 'Invalid athlete id'}, status=status.HTTP_400_BAD_REQUEST)

    coach = get_object_or_404(User, id=coach_id)
    if not coach.is_staff:
        return Response(
            {'message': f'User is not a coach'},
            status=status.HTTP_400_BAD_REQUEST)

    rating = request.data.get('rating', None)
    if not rating:
        return Response({'message': 'rating field is required '}, status=status.HTTP_400_BAD_REQUEST)

    try:
        subscribe = Subscribe.objects.get(subscriber=athlete, subscribed_to=coach)
    except Subscribe.DoesNotExist:
        return Response({'message': 'Athlete must be subscribed to coach'}, status=status.HTTP_400_BAD_REQUEST)

    subscribe.rating = rating
    try:
        subscribe.full_clean()
    except ValidationError as e:
        return Response({'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    subscribe.save()
    return Response({f'message': 'Coach was successfully rated'}, status=status.HTTP_200_OK)
